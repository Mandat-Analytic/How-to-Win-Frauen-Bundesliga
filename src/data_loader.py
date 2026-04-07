import pandas as pd
import numpy as np
import os
import re
import streamlit as st
import openpyxl
from pathlib import Path
from .utils import (
    PLAYER_DATA_DIR, TEAM_DATA_DIR, FOLDER_TO_LEAGUE, 
    LEAGUE_GROUPS, normalize_season, TIER_ORDER
)
# Lazy import/circular dependency handling if needed
# But we made a placeholder so it's fine.
from .feature_engineering import assign_tiers

def discover_files():
    """
    Walks data directories and returns a nested dictionary of file paths.
    Structure: {
        "players": {league: {season: filepath}},
        "teams": {league: {team_name: {season: filepath}}}
    }
    """
    discovered = {"players": {}, "teams": {}}
    
    # --- DISCOVER PLAYERS ---
    # Expected: data/players/<League>/<Season>.xlsx
    if PLAYER_DATA_DIR.exists():
        for league_folder in PLAYER_DATA_DIR.iterdir():
            if league_folder.is_dir():
                # Map folder name to canonical league name
                league_name = FOLDER_TO_LEAGUE.get(league_folder.name, league_folder.name)
                
                if league_name not in discovered["players"]:
                    discovered["players"][league_name] = {}
                
                for file_path in league_folder.glob("*.xlsx"):
                    # Parse season from filename "25_26.xlsx"
                    season_raw = file_path.stem
                    season = normalize_season(season_raw)
                    discovered["players"][league_name][season] = str(file_path)

    # --- DISCOVER TEAMS ---
    # Expected: data/teams/<League>/Team Stats <Name> <Season>.xlsx
    if TEAM_DATA_DIR.exists():
        for league_folder in TEAM_DATA_DIR.iterdir():
            if league_folder.is_dir():
                league_name = FOLDER_TO_LEAGUE.get(league_folder.name, league_folder.name)
                
                if league_name not in discovered["teams"]:
                    discovered["teams"][league_name] = {}
                
                for file_path in league_folder.glob("*.xlsx"):
                    filename = file_path.name
                    # Regex to parse "Team Stats <Name> <Season>.xlsx"
                    # Matches "Team Stats Johor Darul Ta'zim 22_23.xlsx"
                    match = re.search(r"Team Stats (.+) (\d{2}_\d{2})\.xlsx", filename)
                    
                    if match:
                        team_name = match.group(1).strip()
                        season_raw = match.group(2)
                        season = normalize_season(season_raw)
                        
                        if team_name not in discovered["teams"][league_name]:
                            discovered["teams"][league_name][team_name] = {}
                        
                        discovered["teams"][league_name][team_name][season] = str(file_path)
                    else:
                        # Fallback or log warning?
                        print(f"Skipping non-conforming file: {filename}")

    return discovered

def load_team_excel(filepath: str):
    """
    Loads a Wyscout team export Excel file, handling merged cells and alternating rows.
    """
    try:
        # Load workbook with openpyxl to handle merged cells
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        data = []
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Read all rows
        for r in range(1, max_row + 1):
            row_data = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                # Handle merged cells: value is in top-left
                # openpyxl's data_only=True often helps, 
                # but explicit check might be needed if using older versions.
                # Here we assume standard read works or use value.
                val = cell.value
                row_data.append(val)
            data.append(row_data)
        
        # Convert to DataFrame to easier manipulation
        # Assuming row 0 is header (or rows 0-1 are merged headers)
        # Wyscout usually has complex headers. 
        # Strategy: Use pandas read_excel with header detection, forcing merged cells to fill?
        
        # Better approach for speed and robustness with pandas directly:
        # Pandas handles simple merged headers well if we specify header rows.
        # But data is alternating.
        
        df = pd.read_excel(filepath, header=[0]) # Try single header first
        
        # Check for alternating rows
        # Wyscout exports often have "Team" column or similar.
        # If not, use even/odd strategy.
        
        # Identify Team vs Opponent
        # Look for a column that might indicate this.
        # Common Wyscout cols: "Team", "Match", "Date"
        
        # Data Cleaning: 
        # Dropping totally empty rows/cols
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # Detect alternator
        is_team_row = np.zeros(len(df), dtype=bool)
        
        # Heuristic: Even indices (0, 2, 4...) are usually the Team's stats
        # Odd indices (1, 3, 5...) are Opponent stats in that match
        # UNLESS there is a "Team" column specifying name.
        
        indicators = [c for c in df.columns if "team" in str(c).lower()]
        
        if len(df) % 2 != 0:
            # Odd number of rows? Suspicious for strict alternating.
            # But let's assume the heuristic holds for matches.
            pass

        # Apply heuristic
        # row 0 -> Team, row 1 -> Opponent
        team_indices = list(range(0, len(df), 2))
        opp_indices = list(range(1, len(df), 2))
        
        team_df = df.iloc[team_indices].reset_index(drop=True)
        opp_df = df.iloc[opp_indices].reset_index(drop=True)
        
        # Aggregate (Mean)
        # Select numeric columns only
        numeric_cols = team_df.select_dtypes(include=[np.number]).columns
        
        team_stats = team_df[numeric_cols].mean()
        opp_stats = opp_df[numeric_cols].mean()
        
        # Rename opponent stats index to prefix 'opponent_'
        opp_stats.index = ["opponent_" + str(c) for c in opp_stats.index]
        
        return {
            "team_stats": team_stats,
            "opponent_stats": opp_stats,
            "raw_team": team_df,
            "raw_opponent": opp_df
        }

    except Exception as e:
        print(f"Error loading team file {filepath}: {e}")
        return None

def load_player_excel(filepath: str, league: str, season: str):
    """
    Loads player data, filtering for full-backs and ample playing time.
    """
    try:
        df = pd.read_excel(filepath)
        
        # Standardize columns
        df.columns = [c.strip() for c in df.columns]
        
        # Essential columns map (handle variations if any)
        # Wyscout usually has 'Position', 'Minutes played'
        
        if 'Position' not in df.columns:
            # Try finding column with 'Position'
            pass 
            
        # Filter Position
        # Regex: r'\b(RB|RWB|LB|LWB)\b'
        mask_pos = df['Position'].astype(str).apply(lambda x: bool(re.search(r'\b(RB|RWB|LB|LWB)\b', x)))
        df = df[mask_pos]
        
        # Filter Minutes
        if 'Minutes played' in df.columns:
             # Handle "1,200" string format
            if df['Minutes played'].dtype == object:
                 df['Minutes played'] = df['Minutes played'].astype(str).str.replace(',', '').astype(float)
            
            df = df[df['Minutes played'] >= 450]
        
        # Add metadata
        df['league'] = league
        df['season'] = season
        
        # Clean col names
        df.columns = [c.lower().replace(' ', '_').replace('.', '').replace('/', '_') for c in df.columns]
        
        # Normalize 'team' column variations
        # 1. Exact match from known list
        found_team = False
        for col in ['current_team', 'team_name', 'club', 'team_within_selected_timeframe']:
            if col in df.columns:
                df = df.rename(columns={col: 'team'})
                found_team = True
                break
        
        # 2. Fuzzy match if still not found (look for 'team' or 'squad' in column name)
        if not found_team:
            for c in df.columns:
                if 'team' in c or 'squad' in c:
                    df = df.rename(columns={c: 'team'})
                    break

        return df
        
    except Exception as e:
        print(f"Error loading player file {filepath}: {e}")
        return pd.DataFrame()

@st.cache_data
def build_master_datasets(version=1):
    """
    Orchestrates the loading of all data.
    """
    files = discover_files()
    
    # --- BUILD PLAYERS MASTER ---
    player_dfs = []
    for league, seasons in files["players"].items():
        league_group = LEAGUE_GROUPS.get(league, "Other")
        for season, path in seasons.items():
            df = load_player_excel(path, league, season)
            if not df.empty:
                df['league_group'] = league_group
                player_dfs.append(df)
    
    if player_dfs:
        players_master = pd.concat(player_dfs, ignore_index=True)
    else:
        players_master = pd.DataFrame()

    # --- BUILD TEAMS MASTER ---
    team_rows = []
    for league, teams in files["teams"].items():
        league_group = LEAGUE_GROUPS.get(league, "Other")
        for team, seasons in teams.items():
            for season, path in seasons.items():
                res = load_team_excel(path)
                if res:
                    # Combine team and opponent stats into one row
                    row = pd.concat([res["team_stats"], res["opponent_stats"]])
                    row['team'] = team
                    row['league'] = league
                    row['league_group'] = league_group
                    row['season'] = season
                    team_rows.append(row)
    
    if team_rows:
        teams_master = pd.DataFrame(team_rows)
    else:
        teams_master = pd.DataFrame()

    # --- ASSIGN TIERS ---
    if not teams_master.empty:
        # We need to assign tiers based on logic in feature_engineering
        # teams_master matches the input expectations?
        teams_master = assign_tiers(teams_master)
        
        # Also assign tiers to players based on their team? 
        # Player export has "Team" column usually.
        # We need to map team tiers to players.
        if not players_master.empty:
            if 'team' in players_master.columns:
                # Create a lookup
                # Unique team-season-league tier
                tier_lookup = teams_master[['team', 'season', 'tier']].drop_duplicates()
                # Merge? or Map?
                # Players 'team' column might differ slightly in spelling from Team Stats file.
                # For now, simplistic merge.
                players_master = players_master.merge(tier_lookup, on=['team', 'season'], how='left')
    
    # Ensure tier exists even if merge failed or teams empty
    if not players_master.empty and 'tier' not in players_master.columns:
        players_master['tier'] = "Mid" # Default
    else:
        # Fill missing tiers from partial merge
        if not players_master.empty:
            players_master['tier'] = players_master['tier'].fillna("Mid")

    return teams_master, players_master
